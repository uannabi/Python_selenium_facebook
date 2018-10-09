import openpyxl
import xlwt
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from getpass import getpass
import time
import codecs
import sys
import os

UTF8Writer = codecs.getwriter('utf8')
sys.stdout = UTF8Writer(sys.stdout)

usr = input('Enter your username or email id: ') #enter mail like 'name@mail.com' in console
pwd = getpass('enter your password: ') #enter passward in console


# Specifying incognito mode as you launch your browser[OPTIONAL]
option = webdriver.ChromeOptions()
option.add_argument("--incognito")

# Create new Instance of Chrome in incognito mode
driver = webdriver.Chrome(chrome_options=option)

# driver = webdriver.ChromeOptions()
driver.get('https://www.facebook.com/')  # facebook link



username_box = driver.find_element_by_id('email')
username_box.send_keys(usr)

password_box = driver.find_element_by_id('pass')
password_box.send_keys(pwd)

login_btn = driver.find_element_by_id('u_0_2')
login_btn.submit()
driver.get('https://www.facebook.com/groups/cookupsBD/')  # facebook group link
time.sleep(50); #sleep time for load all post

wb = Workbook()
filepath = "output.xlsx" #output file name
sheet = wb.active #open xl file
# ani = 13


def writingToExcel(label, listItems, columNumber):
    sheet.cell(row=1, column=columNumber).value = label
    i = 2
    while i < (len(listItems)):
        sheet.cell(row=i, column=columNumber).value = listItems[i - 2]
        i += 1


coock_name_junk = driver.find_elements_by_xpath("//span[@class='fwn fcg']//a[@class='profileLink'][text()]")
coock = [x.text for x in coock_name_junk]  # coock name



product_name_junk = driver.find_elements_by_xpath("//div//div[@class='_l53']//span[position()=2][text()]")
products = [x.text for x in product_name_junk]  # same concept as for-loop/ list-comprehension above.



product_price_junk = driver.find_elements_by_xpath("//div//div[@class='_l57'][text()]")
prices = [x.text for x in product_price_junk]  # same concept as for-loop/ list-comprehension above.

product_location_junk = driver.find_elements_by_xpath("//div//div[@class='_l56']//div[position()=2]")
locations = [x.text for x in product_location_junk]  # same concept as for-loop/ list-comprehension above.

product_post_time_junk = driver.find_elements_by_xpath("//span[@class='fsm fwn fcg']//a")
times = [x.text for x in product_post_time_junk]  # same concept as for-loop/ list-comprehension above.

product_description_junk = driver.find_elements_by_xpath(
    "//div[@class='_5pbx userContent _3576']//div//div//p[position()=1]")
descriptions = [x.text for x in product_description_junk]  # same concept as for-loop/ list-comprehension above.

# product_photo_junk = driver.find_elements_by_xpath("//div[@class='_6l- __c_']//div//img//@src")
order_link_junk = driver.find_elements_by_xpath("//div[@class='_5pbx userContent _3576']//div//p[position()=2]//a")
link = [x.text for x in order_link_junk]  # same concept as for-loop/ list-comprehension above.



writingToExcel('Cook Name', coock, 1)
writingToExcel('Time of Post', times, 2)
writingToExcel('Title of Post', products, 3)
writingToExcel('Cook Location', locations, 4)
writingToExcel('Price', prices, 5)
writingToExcel('Description', descriptions, 6)
writingToExcel('Order Link', link, 7)
# writingToExcel('Image Link',imglink, 8)
wb.save(filepath) #save xl file

